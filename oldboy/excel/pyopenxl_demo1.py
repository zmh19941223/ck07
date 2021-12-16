from openpyxl import Workbook,load_workbook

# wb = Workbook()
#
# sheet = wb.active  # 获取当前激活的sheet
#
# print(sheet.title)
#
# # 修改sheet名称
# sheet.title = '朱明浩的后宫们'
#
# sheet['B3'] = '李露露'
# sheet.append([1,2,3])  # 直接在下一行的第一个单元格一次写入数据
#
# wb.save('test.xlsx')


# 载入已存在的Excel文件

wb = load_workbook("test.xlsx")

print(wb.sheetnames)
# print(wb.get_sheet_by_name('Sheet1'))

sheet = wb.get_sheet_by_name('Sheet1')
B3 = sheet['B3']  # 获取单元格

# for cell in sheet["B3:B6"]:
#     print(cell[0].value)


# 循环遍历每一行
# for row in sheet:
#     for cell in row:
#         print(cell.value, end='\t')
#     print()
# print(B3.value)

# 遍历表的一部分区域
# for row in sheet.iter_rows(min_col=3, max_row=10, max_col=5):
#     for cell in row:
#         print(cell.value, end=',')
#     print()

# 按列遍历
for col in sheet.columns:
    # print(col)
    for cell in col:
        print(cell.value, end=',')

    print()