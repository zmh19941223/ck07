"""
============================
Author:柠檬班-木森
Time:2020/12/5  10:40
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""

import openpyxl



class HandleExcel:

    def __init__(self, filename, sheetname):
        """
        :param filename: elcle文件名(路径)
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """读取excel数据"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        # 获取第一行的表头
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行之外的其他行
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        # 返回读取出来的数据
        return cases

    def write_data(self, row, column, value):
        """
        数据写入的方法
        :param row: 写入的行
        :param column: 写入的列
        :param value: 写入的值
        :return:
        """
        # 加载工作簿对象
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        # 写入数据
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)
