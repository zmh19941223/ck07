"""
============================
Author:柠檬班-木森
Time:2020/12/3  21:30
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""

import openpyxl

# -----------------基础用法------------------------------------
# 1、加载excel文件为工作簿对象
# workbook = openpyxl.load_workbook('test001.xlsx')
# 获取所有的表单名
# print(workbook.sheetnames)
# 2、选中表单
# sh = workbook["Sheet1"]
# print(sh)

# 3、读取数据
# c = sh.cell(row=1,column=1)
# print(c.value)
# print(sh.cell(row=3,column=3).value)


# ------------------读取表单中所有的数据-------------------------
workbook = openpyxl.load_workbook('test001.xlsx')
sh = workbook["Sheet1"]

# rows:按行获取表单中所有的格子,每一行的格子放到一个元组中
# res = list(sh.rows)
# for i in res:
#     print(i)


# rows:按行获取表单中所有的格子,每一行的格子放到一个元组中
# res = list(sh.columns)
# for i in res:
#     print(i)
